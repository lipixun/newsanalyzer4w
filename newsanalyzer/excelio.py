# encoding=utf8

""" The input and output
    Author: lipixun
    Created Time : 日  2/12 14:15:18 2017

    File Name: io.py
    Description:

"""

import logging

from csv import DictWriter
from openpyxl import load_workbook

from .spec import SheetCountry, SheetRegion, SheetProvince, SheetCity, SheetNews
from .model import Country, Region, Province, City, News

class ExcelInput(object):
    """The excel input
    """
    logger = logging.getLogger("newsanalyzer.ExcelInput")

    def __init__(self, filename):
        """Create a new ExcelInput
        """
        # Initialize
        self.countries = []
        self.regions = []
        self.provinces = []
        self.cities = []
        self.news = []
        # Load excel
        self.workbook = load_workbook(filename)
        # Get data from workbook
        # Country
        self.logger.info("Load country")
        countrySheet = self.workbook[SheetCountry]
        self.loadCountryFromSheet(countrySheet)
        self.logger.info("Totally load [%d] countries", len(self.countries))
        if self.logger.isEnabledFor(logging.DEBUG):
            for country in self.countries:
                self.logger.debug("Loaded country: %s", country)
        # Region
        self.logger.info("Load region")
        regionSheet = self.workbook[SheetRegion]
        self.loadRegionFromSheet(regionSheet)
        self.logger.info("Totally load [%d] regions", len(self.regions))
        if self.logger.isEnabledFor(logging.DEBUG):
            for region in self.regions:
                self.logger.debug("Loaded region: %s", region)
        # Province
        self.logger.info("Load province")
        provinceSheet = self.workbook[SheetProvince]
        self.loadProvinceFromSheet(provinceSheet)
        self.logger.info("Totally load [%d] provinces", len(self.provinces))
        if self.logger.isEnabledFor(logging.DEBUG):
            for province in self.provinces:
                self.logger.debug("Loaded province: %s", province)
        # City
        self.logger.info("Load city")
        citySheet = self.workbook[SheetCity]
        self.loadCityFromSheet(citySheet)
        self.logger.info("Totally load [%d] cities", len(self.cities))
        if self.logger.isEnabledFor(logging.DEBUG):
            for city in self.cities:
                self.logger.debug("Loaded city: %s", city)
        # People
        # News
        self.logger.info("Load news")
        newsSheet = self.workbook[SheetNews]
        self.loadNewsFromSheet(newsSheet)
        self.logger.info("Totally load [%d] news", len(self.news))
        if self.logger.isEnabledFor(logging.DEBUG):
            for news in self.news:
                self.logger.debug("Loaded news: %s", news)

    def normalize(self, text):
        """Normalize text
        """
        if not text:
            return
        # A tricky way to remove useless spaces
        return text.replace("  ", " ").replace("  ", " ").replace("  ", " ").replace("  ", " ").lower()

    def loadCountryFromSheet(self, sheet):
        """Load country from sheet
        """
        for i in range(sheet.min_row, sheet.max_row + 1):
            name = self.normalize(sheet.cell(row = i, column = 1).value)
            if not name:
                raise ValueError("Country name must not be empty")
            alias = [ name ]
            index = 1
            while True:
                index += 1
                aliasName = self.normalize(sheet.cell(row = i, column = index).value)
                if not aliasName:
                    break
                alias.append(aliasName)
            # Add country
            self.countries.append(Country(name, alias))

    def loadRegionFromSheet(self, sheet):
        """Load region from sheet
        """
        # Read name and alias
        index = 0
        while True:
            index += 1
            name = self.normalize(sheet.cell(row = 1, column = index).value)
            alias = self.normalize(sheet.cell(row = 2, column = index).value)
            if not name or not alias:
                break
            self.regions.append(Region(name, [ alias ]))
        # Read countries
        for i in range(sheet.min_row + 2, sheet.max_row + 1):
            index = 0
            while True:
                index += 1
                country = self.normalize(sheet.cell(row = i, column = index).value)
                if not country:
                    break
                self.regions[index - 1].countries.append(country.strip())

    def loadProvinceFromSheet(self, sheet):
        """Load province from sheet
        """
        for i in range(sheet.min_row, sheet.max_row + 1):
            name = self.normalize(sheet.cell(row = i, column = 1).value)
            if not name:
                raise ValueError("Province name must not be empty")
            alias = [ name ]
            index = 1
            while True:
                index += 1
                aliasName = self.normalize(sheet.cell(row = i, column = index).value)
                if not aliasName:
                    break
                alias.append(aliasName)
            # Add country
            self.provinces.append(Province(name, alias))

    def loadCityFromSheet(self, sheet):
        """Load country from sheet
        """
        for i in range(sheet.min_row, sheet.max_row + 1):
            name = self.normalize(sheet.cell(row = i, column = 1).value)
            if not name:
                raise ValueError("City name must not be empty")
            alias = [ name ]
            index = 1
            while True:
                index += 1
                aliasName = self.normalize(sheet.cell(row = i, column = index).value)
                if not aliasName:
                    break
                alias.append(aliasName)
            # Add country
            self.cities.append(City(name, alias))

    def loadNewsFromSheet(self, sheet):
        """Load news from sheet
        """
        for i in range(sheet.min_row + 2, sheet.max_row + 1):
            title = self.normalize(sheet.cell(row = i, column = 6).value)
            content = self.normalize(sheet.cell(row = i, column = 7).value)
            if title and content:
                self.news.append(News(title.strip(), content.strip()))

class KeywordResultWriter(object):
    """The keyword result writer
    """
    FieldKeyword    = u"关键词"
    FieldTermsCount = u"单词个数"
    FieldTFIDF      = u"TF-IDF"
    FieldFrequency  = u"出现次数"

    def __init__(self, outStream):
        """Create a new KeywordResultWriter
        """
        self.writer = DictWriter(outStream, fieldnames=[
            self.FieldKeyword,
            self.FieldTermsCount,
            self.FieldTFIDF,
            self.FieldFrequency,
            ])
        self.writer.writeheader()

    def write(self, data):
        """Write data
        """
        self.writer.writerow(data)

class CooccurrenceResultWriter(object):
    """The cooccurrence result writer
    """
    FieldKeyword    = u"关键词"
    FieldCoword     = u"相关词汇"
    FieldTermsCount = u"单词个数"
    FieldTFIDF      = u"TF-IDF"
    FieldFrequency  = u"出现次数"

    def __init__(self, outStream):
        """Create a new KeywordResultWriter
        """
        self.writer = DictWriter(outStream, fieldnames=[
            self.FieldKeyword,
            self.FieldCoword,
            self.FieldTermsCount,
            self.FieldTFIDF,
            self.FieldFrequency,
            ])
        self.writer.writeheader()

    def write(self, data):
        """Write data
        """
        self.writer.writerow(data)

class CooccurrenceEntityResultWriter(object):
    """The cooccurrence result writer
    """
    FieldKeyword    = u"关键词"
    FieldEntityType = u"相关实体类型"
    FieldEntity     = u"相关实体"
    FieldFrequency  = u"出现次数"

    def __init__(self, outStream):
        """Create a new KeywordResultWriter
        """
        self.writer = DictWriter(outStream, fieldnames=[
            self.FieldKeyword,
            self.FieldEntityType,
            self.FieldEntity,
            self.FieldFrequency,
            ])
        self.writer.writeheader()

    def write(self, data):
        """Write data
        """
        self.writer.writerow(data)
